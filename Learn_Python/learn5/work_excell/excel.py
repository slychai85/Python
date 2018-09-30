import csv
from openpyxl import Workbook

def read_csv( filename ):
    # Список куда будем добавлять словари
    data =[]

    # Открываем файл и создаем словари из него
    with open(filename, 'r', encoding='utf-8') as f:
        fields = [ 'Участник проекта', 'Страна', 'Павильон, зал',	'Стенд', 'Тематический раздел' ]
        reader = csv.DictReader(f, fields, delimiter='\t')
        for row in reader:
            data.append( row )

    return data


def excel_write( data ):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Test'

    worksheet.cell(row=1, column=1).value = 'Участник проекта'
    worksheet.cell(row=1, column=2).value = 'Страна'
    worksheet.cell(row=1, column=3).value = 'Павильон, зал'
    worksheet.cell(row=1, column=4).value = 'Стенд'
    worksheet.cell(row=1, column=5).value = 'Тематический раздел'

    row = 2

    for item in data:
        worksheet.cell(row=row, column=1).value = item['Участник проекта']
        worksheet.cell(row=row, column=2).value = item['Страна']
        worksheet.cell(row=row, column=3).value = item['Павильон, зал']
        worksheet.cell(row=row, column=4).value = item['Стенд']
        worksheet.cell(row=row, column=5).value = item['Тематический раздел']

        row += 1

    workbook.save( 'test_excel.xlsx' )

csv_data = read_csv('test_excel.csv' )
